import numpy as np
from tkinter import messagebox, filedialog
import openpyxl as pxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import re
import matplotlib.pyplot as plt
import pandas as pd
import xlwings as xw
import traceback
import os
from dataclasses import dataclass

# This class is intended to synchronize WellData and corresponding DrugReports metadata
class AssayMetadata:
    date = []
    ctr = []
    plateNo = []
    drug = []
    receptor = []
    concentration = []
    h5Path = []
    rawDataPath = []

    def __init__(self):
        pass

    def display(self):
        print(self.date)
        print(self.ctr)
        print(self.receptor)
        print(self.h5Path)
        print(self.rawDataPath)

# Stores 96 well-plate pharmacology assay data. Base class for multiple assays
class WellData:
    metadata = AssayMetadata()
    comments = []               # list
    data = []                   # well plate data
    omittedVals = []            # list of values from original well data that are omitted by user selection

    def __init__(self, df, plateNum=1):
        if df.size == 96:
            self.data = df
        else:
            self.data = df.loc[df.index[0:8], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]]
        self.metadata.plateNo = plateNum

    def display(self):
        print(self.data)
        for c in self.comments:
            print(c)

    def add_comment(self, comment):
        self.comments.append(comment)

    def set_date(self, dateStr):
        self.metadata.date = dateStr

# For use with binding assays 
class BindingPlate(WellData):
    drugs = []                  # list of drugs on plate, one per two rows
    totals = []                 # well plate totals with no drug
    nsb = []                    # well plate non-specific binding
    highestConc = []            # log [drug]
    specificActivityCi = []     # concentration entered by user
    specificActivityCpm = []    # concentration converted  
    volMl = []                  # volume of radioactive ligand added

    def __init__(self, df, plate):
        super().__init__(df, plate)
        self.drugs = ["Drug1", "Drug2", "Drug3", "Drug4"]
        self.highestConc = [None, None, None, None]
        plateTotals = self.data.loc[self.data.index[0:8], [1, 12]]
        self.totals = np.average(plateTotals)
        plateNSB = self.data.loc[self.data.index[0:8], 2]
        self.nsb = np.average(plateNSB) 

    def display(self):
        super().display()
        for d in self.drugs:
            print(d)

    def make_drug_report(self, index):
        return DrugReports(self, index)

    def make_all_drug_reports(self):
        reports = []
        for i in range(4):
            reports.append(DrugReports(self, i))
        return reports
    
    def update_drugs(self, drugName, idx):
        self.drugs[idx] = drugName

    def update_conc(self, highestConc, idx):
        self.highestConc[idx] = highestConc

    def update_receptor(self, receptorName):
        self.metadata.receptor = receptorName

# For use with screening assays
class ScreeningPlate(WellData):
    def __init__(self, df, plate):
        super().__init__(df, plate)
        self.receptors = []
        self.drugDict = {}
        self.concDict = {}
        self.experimentIdx = {}

    def display(self):
        super().display()
        
    def make_triplicate_reports(self):
        pass

# Stores assay for one drug/receptor/date combination
class DrugReports:
    metadata = AssayMetadata()  
    drug = []                   # Drug name
    average = []                # Average counts at a given concentration
    specific = []               # Average counts - nonspecific binding
    concentration = []          # Drug concentration in Molarity
    logConc = []                # log [drug]
    pctTotal = []               # specific activity/specific bound
    averageSpecific = []        # mean 
    specificBound = []
    pctSpecificBinding = []
    omittedVals = []

    def __init__(self, wellData, index):
        self.drug = wellData.drugs[index]
        self.metadata = wellData.metadata
        startRow = (index) * 2
        endRow = startRow + 2
        drugData = wellData.data.iloc[startRow:endRow]
        self.averageSpecific = np.average(wellData.totals)
        testData = drugData.loc[drugData.index[:], [3, 4, 5, 6, 7, 8, 9, 10, 11]].values
        self.average = [testData.mean(axis=0)]
        self.specific = [x - wellData.nsb for x in self.average]
        self.specificBound = wellData.totals - wellData.nsb
        self.pctTotal = [x * 100 / self.specificBound for x in self.specific]
        concentrationSteps = np.linspace(0, -4, num=9, endpoint=True)
        self.logConc = concentrationSteps + wellData.highestConc[index]
        concCalc = lambda x: 10 ** x
        self.concentration = concCalc(self.logConc)
        self.average = self.average[0].tolist()
        self.specific = self.specific[0].tolist()
        self.pctTotal = self.pctTotal[0].tolist()

@dataclass
class ScreeningTriplet:
    receptor: str = None
    drug: str = None
    concentration: int = None
    weighing: str = None
    plate: int = None

# Unsure where this class is going - potentially a simple aggregator that calculates % inhib for screening assays
class TriplicateScreen:
    def __init__(self, vals, metadata, nsb, totals):
        self.values = vals
        self.metadata = metadata
        self.nsb = np.mean(nsb)
        self.totals = np.mean(totals)

    def calculate_totals(self):
        specificActivity = np.mean(self.values) - self.nsb
        totalSpecific = np.mean(self.totals) - self.nsb
        
# Produces summary tables for NIDA reports
class SummaryTable:
    receptors = ["D1", "D2", "D3", "D4", 
                 "1A", "2A", "2B", "2C"]
    bindingColumns = ["IC50", "Ki", "Hill Slope"]
    functionColumns = ["Agonist EC50", "Standard Agonist", "Percent Stimulation", 
                       "Antagonist IC50", "Standard Antagonist", "Percent Inhibition"]
    srcDir = []
    summary = {}

    def __init__(self, dir):
            self.srcDir = dir
            self.import_binding_data()
            self.import_function_data()

    # the property 'summary' is used to build all the summary tables, it stores information scraped from xlsx
    def add_receptor(self, drugID, receptor):
        if drugID not in self.summary:
            self.summary[drugID] = {}
        if receptor not in self.summary[drugID]:
            self.summary[drugID][receptor] = {
                "mean": {
                    "binding":{"ic50": None, "ki": None, "hillSlope": None},
                    "function":{"ec50": None, "pctStim": None, "ic50": None, "pctInhib": None}
                    },
                "sem": {
                    "binding":{"ic50": None, "ki": None, "hillSlope": None},
                    "function":{"ec50": None, "pctStim": None, "ic50": None, "pctInhib": None}
                    }
            }

    # identifies the columns within a particular row that contain the given header
    def find_excel_header(self, row, header):
        indices = [i for i, item in enumerate(row) if isinstance(item, str) and re.search(header, item, re.IGNORECASE)]
        return indices
    
    # extracts the name of the drug from the first column for a given group.
    def find_drug_name(self, sheet, idx):
        drugID = sheet.cell(row = idx+1, column=1).value
        if not drugID:
            temp = idx
        while not drugID:
            temp = temp-1
            drugID = sheet.cell(row = temp+1, column=1).value
        drugID = str(drugID).splitlines()
        return drugID[0]

    # uses string format magic to round a number to a string based on the number of significant figures
    def round_sig(self, x, sigFigs):
        if not x or x=='#DIV/0!' or x=='#VALUE!':
            return '0'
        rounded = '{:g}'.format(float('{:.{p}g}'.format(x, p=sigFigs)))
        if len(rounded) < sigFigs and '.' not in rounded:
            rounded = rounded + '.'
        while len(rounded) <= sigFigs and '.' in rounded:
            rounded = rounded + '0'
        return rounded
    
    # allows users to select which files contain binding data
    def import_binding_data(self):
        loadMorePrompt = True
        while loadMorePrompt:
            fileName = filedialog.askopenfilename(initialdir=self.srcDir,
                                                  title="Select binding data: ")
            workbook = pxl.load_workbook(fileName, data_only=True)
            self.parse_binding_table(workbook)
            loadMorePrompt = messagebox.askyesno("Load more binding data?")

    # makes sure sheets are named according to allowed receptors (deprecated if using TemplateGenerator)
    def parse_binding_table(self, workbook):
        for name in workbook.sheetnames:
            matchingReceptor = [isinstance(item, str) and re.search(item, name, re.IGNORECASE) for item in self.receptors]
            if not any(matchingReceptor):
                print(f"No matching receptor found for workbook page {name}")
                continue
            receptorName = [self.receptors[i] for i, match in enumerate(matchingReceptor) if match]
            sheet = workbook[name]
            receptorName = receptorName[0]
            if type(receptorName) is not str:
                receptorName = str(receptorName)
            self.parse_binding_summary_sheet(sheet, receptorName)
        
    # goes through a binding sheet to extract data for summary table
    def parse_binding_summary_sheet(self, sheet, receptor):
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            ic50 = self.find_excel_header(row, "ic50")
            ki = self.find_excel_header(row, "ki")
            hillSlope = self.find_excel_header(row, "hill slope")
            if ic50 and ki and hillSlope:
                drugID = self.find_drug_name(sheet, idx)
                self.add_receptor(drugID, receptor)
                # relative average and sem indexing to the cells with the datatype's header
                ic50Average = sheet.cell(row=idx+2, column=ic50[0]+2).value
                ic50SEM = sheet.cell(row=idx+2, column=ic50[0]+3).value
                kiAverage = sheet.cell(row=idx+2, column=ki[0]+2).value
                kiSEM = sheet.cell(row=idx+2, column=ki[0]+3).value
                hillAverage = sheet.cell(row=idx+2, column=hillSlope[0]+2).value
                hillSEM = sheet.cell(row=idx+2, column=hillSlope[0]+3).value
                self.summary[drugID][receptor]["mean"]["binding"]["ic50"] = ic50Average
                self.summary[drugID][receptor]["sem"]["binding"]["ic50"] = ic50SEM
                self.summary[drugID][receptor]["mean"]["binding"]["ki"] = kiAverage
                self.summary[drugID][receptor]["sem"]["binding"]["ki"] = kiSEM
                self.summary[drugID][receptor]["mean"]["binding"]["hillSlope"] = hillAverage
                self.summary[drugID][receptor]["sem"]["binding"]["hillSlope"] = hillSEM

    # produces summary table for scraped binding data after parsing excel
    def make_binding_summary_tables(self):
        meanPrecision = 3
        semPrecision = 2
        for drug, nestedDict in self.summary.items():
            print(f"Drug: {drug}")
            df = pd.DataFrame()
            df['Receptor'] = list(self.summary[drug].keys())
            ic50 = []
            ki = []
            hillSlope = []
            for subKey, values in nestedDict.items():
                if isinstance(values, dict):
                    meanVals = values['mean']['binding']
                    semVals = values['sem']['binding']
                    combinedVals = {}
                    for meanKey, semKey in zip(meanVals.keys(), semVals.keys()):
                        outputStr = f"{self.round_sig(meanVals[meanKey], meanPrecision)} ± {self.round_sig(semVals[semKey], semPrecision)}"
                        match(meanKey):
                            case('ic50'):
                                ic50.append(outputStr)
                            case('ki'):
                                ki.append(outputStr)
                            case('hillSlope'):
                                hillSlope.append(outputStr)
            df['IC₅₀ (nM) ± SEM'] = ic50
            df['Ki (nM) ± SEM'] = ki
            df['Hill Slope ± SEM'] = hillSlope
            fig, ax = plt.subplots()
            table = ax.table(cellText=df.values,
                        colLabels=df.columns,
                        loc='center',
                        cellLoc='center')
            table = self.format_table(table, len(df.columns))
            ax.set_axis_off()
            plt.tight_layout()
            plt.savefig(f'{self.srcDir}\{drug}_binding_table.png', bbox_inches='tight')
            plt.close(fig)
            
    # allows users to specify which files contain function data
    def import_function_data(self):
        loadMorePrompt = True
        while loadMorePrompt:
            fileName = filedialog.askopenfilename(initialdir=self.srcDir,
                                                  title="Select function data: ")
            workbook = pxl.load_workbook(fileName, data_only=True)
            workbook.name = fileName
            self.parse_function_table(workbook)
            loadMorePrompt = messagebox.askyesno("Load more function data?")

    # makes sure sheets are named according to allowed receptors (deprecated if using TemplateGenerator)
    def parse_function_table(self, workbook):
        matchingFilename = [isinstance(item, str) and re.search(item, workbook.name, re.IGNORECASE) for item in self.receptors]
        for name in workbook.sheetnames:
            print(name)
            matchingReceptor = [isinstance(item, str) and re.search(item, name, re.IGNORECASE) for item in self.receptors]
            if not any(matchingReceptor) and not any(matchingFilename):
                print(f"No matching receptor found for workbook page {name}")
                continue
            if any(matchingFilename) and sum(1 for item in matchingFilename if item is not None):
                receptorName = [self.receptors[i] for i, match in enumerate(matchingFilename) if match]
            elif any(matchingReceptor):
                receptorName = [self.receptors[i] for i, match in enumerate(matchingReceptor) if match]
            print(receptorName)
            sheet = workbook[name]
            receptorName = receptorName[0]
            if type(receptorName) is not str:
                receptorName = str(receptorName)
            print(receptorName)
            self.parse_function_summary_sheet(sheet, receptorName)

    # goes through a function sheet to extract data for summary table
    def parse_function_summary_sheet(self, sheet, receptor):
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            ec50 = self.find_excel_header(row, "ec50")
            ic50 = self.find_excel_header(row, "ic50")
            pctVal = self.find_excel_header(row, "%")
            ave = self.find_excel_header(row, "mean")
            aveAlt = self.find_excel_header(row, "ave")
            sem = self.find_excel_header(row, "sem")
            if (ec50 or ic50) and (ave or aveAlt) and sem:
                drugID = self.find_drug_name(sheet, idx)
                self.add_receptor(drugID, receptor)
            if ec50 and (ave or aveAlt) and sem:
                # relative average and sem indexing to the cells with the datatype's header
                ec50Average = sheet.cell(row=idx+2, column=ec50[0]+2).value
                ec50SEM = sheet.cell(row=idx+2, column=ec50[0]+3).value
                pctStimAverage = sheet.cell(row=idx+2, column=pctVal[0]+2).value
                pctStimSEM = sheet.cell(row=idx+2, column=pctVal[0]+3).value
                self.summary[drugID][receptor]["mean"]["function"]["ec50"] = ec50Average
                self.summary[drugID][receptor]["sem"]["function"]["ec50"] = ec50SEM
                self.summary[drugID][receptor]["mean"]["function"]["pctStim"] = pctStimAverage
                self.summary[drugID][receptor]["sem"]["function"]["pctStim"] = pctStimSEM
            elif ic50 and (ave or aveAlt) and sem:
                # relative average and sem indexing to the cells with the datatype's header
                ic50Average = sheet.cell(row=idx+2, column=ic50[0]+2).value
                ic50SEM = sheet.cell(row=idx+2, column=ic50[0]+3).value
                pctInhibAverage = sheet.cell(row=idx+2, column=pctVal[0]+2).value
                pctInhibSEM = sheet.cell(row=idx+2, column=pctVal[0]+3).value
                self.summary[drugID][receptor]["mean"]["function"]["ic50"] = ic50Average
                self.summary[drugID][receptor]["sem"]["function"]["ic50"] = ic50SEM
                self.summary[drugID][receptor]["mean"]["function"]["pctInhib"] = pctInhibAverage
                self.summary[drugID][receptor]["sem"]["function"]["pctInhib"] = pctInhibSEM

    # produces summary table for scraped function data after parsing excel
    def make_function_summary_tables(self):
        meanPrecision = 3
        semPrecision = 2
        for drug, nestedDict in self.summary.items():
            df = pd.DataFrame()
            df['Receptor'] = list(self.summary[drug].keys())
            ec50 = []
            pctStim = []
            ic50 = []
            pctInhib = []
            for subKey, values in nestedDict.items():
                if isinstance(values, dict):
                    meanVals = values['mean']['function']
                    semVals = values['sem']['function']
                    combinedVals = {}
                    for meanKey, semKey in zip(meanVals.keys(), semVals.keys()):
                        outputStr = f"{self.round_sig(meanVals[meanKey], meanPrecision)} ± {self.round_sig(semVals[semKey], semPrecision)}"
                        match(meanKey):
                            case('ec50'):
                                ec50.append(outputStr)
                            case('pctStim'):
                                pctStim.append(outputStr)
                            case('ic50'):
                                ic50.append(outputStr)
                            case('pctInhib'):
                                pctInhib.append(outputStr)
            df['Agonist EC₅₀\n(nM) ± SEM'] = ec50
            df['% Stimulation'] = pctStim
            df['Antagonist IC₅₀\n(nM) ± SEM'] = ic50
            df['% Inhibition'] = pctInhib
            fig, ax = plt.subplots()
            table = ax.table(cellText=df.values,
                        colLabels=df.columns,
                        loc='center',
                        cellLoc='center')
            table = self.format_table(table, len(df.columns))
            ax.set_axis_off()
            plt.tight_layout()
            plt.savefig(f'{self.srcDir}\{drug}_function_table.png', bbox_inches='tight')
            plt.close(fig)

    def format_table(self, table, columns):
        table.auto_set_font_size(False)
        table.set_fontsize(12)
        table.auto_set_column_width(col=list(range(columns)))
        for r in range(0, columns):
            cell = table[0, r]
            cell.set_height(0.13)
        return table
    
# Produces blank excel spreadsheets to be later used by SummaryTable objects
class TemplateGenerator:
    def __init__(self, 
                 saveDir=r'E:/PharmOps-sample-data/summary table/', 
                 drugNames=[],
                 standardsDict={"binding": {}, "function": {}}
                 ):
        self.saveDir = saveDir
        if self.saveDir[-1] != "/":
            self.saveDir += "/"
        self.drugNames = drugNames
        self.standards = standardsDict
        self.blankRows = 2                  
        self.bindingDA = pxl.Workbook()
        self.binding5HT = pxl.Workbook()
        self.functionDA = pxl.Workbook()
        self.function5HT = pxl.Workbook()   
        self.receptorsDA = ['D1', 'D2', 'D3', 'D4'] 
        self.receptors5HT = ['5HT1A', '5HT2A', '5HT2B', '5HT2C'] 
        self.sheetTemplate = ['Assay Title', '', 'Receptor', '', 'Passage #:', '']
        self.bindingTemplate = ['', 'Date', 
                        'IC50', 'Mean', 'SEM', 
                        '[radioligand]', 'Ki', 'Mean', 'SEM', 
                        'Hill slope', 'Mean', 'SEM', 
                        'Initials/Special Conditions']
        self.agonistTemplate = ['', 'Date', 
                           'EC50', 'Mean', 'SEM', 
                           'Max effect', 'Max standard', 
                           '% max standard', 'Mean', 'SEM', 
                           'Notes']
        self.antagonistTemplate = ['', 'Date', 
                              'IC50', 'Mean', 'SEM',
                              'Top of curve', 'Standard Pct Agonist',
                              '% reversal', 'Mean', 'SEM',
                              'Notes']
        self.bindingRow = ['Assay description', '', 'Radioligand']
        self.functionRow = ['Assay description']
        self.bindingKdCell = "$H$1"
        self.ligandConcCol = 'F'
        self.maxEffectCol = 'F'
        self.standardCol = 'G'
        self.startingRow = 4
        self.make_binding_template("DA")
        self.make_binding_template("5HT")
        self.make_function_template("DA")
        self.make_function_template("5HT")


    def add_binding_section(self, sheet, idx):
        self.startRange = self.startingRow + 1
        self.endRange = self.startingRow + self.blankRows + 1
        sheet.append(self.bindingTemplate)
        self.format_cell(sheet, self.startingRow, range(1, 16), 'Header')
        sheet.cell(row=sheet.max_row, column=1, value=idx)
        sheet.append([])
        sheet.cell(row=sheet.max_row+1, column=7, 
                value=self.add_formula('Ki', 'C', sheet.max_row+1))
        self.format_cell(sheet, self.startingRow+1, [3, 6, 7, 10], 'Table')
        for i in range(0, self.blankRows-1):
            sheet.append([])
            self.format_cell(sheet, self.startingRow+2+i, [3, 6, 7, 10], 'Table')
        sheet.append([])
        self.populate_value_headers(sheet, self.startingRow+1, ['C', 'G', 'J'])
        self.startingRow = self.startingRow + self.blankRows + 2

    def add_standards_header(self, sheet):
        standardFont = Font(name='Arial',
                            size=12,
                            bold=True,
                            underline='single')
        sheet.append(['Standards'])
        cell = sheet.cell(row=sheet.max_row, column=1)
        cell.font = standardFont
        sheet.append([])
        self.startingRow += 2

    # Can produce binding templates for 5HT or DA
    def make_binding_template(self, receptor):
        match(receptor):
            case("DA"):
                wb = self.bindingDA
                receptorList = self.receptorsDA
            case("5HT"):
                wb = self.binding5HT
                receptorList = self.receptors5HT
        ws1 = wb.active
        ws1.title = receptorList[0]
        for i in range(1, 4):
            wb.create_sheet(receptorList[i])
        self.currentWB = wb
        for sheetName in receptorList:
            self.startingRow = 4
            sheet = wb[sheetName]
            sheet.append(self.sheetTemplate)
            self.format_cell(sheet, 1, [1, 3, 5, 7], 'Title')
            sheet.cell(row=1, column=7, value = 'Kd(nM)=')
            sheet.append(self.bindingRow)
            self.format_cell(sheet, 2, [1, 3], 'Title')
            sheet.append([])
            for i in self.drugNames:
                self.add_binding_section(sheet, i)
            if sheetName not in self.standards["binding"]:
                continue
            self.add_standards_header(sheet)
            for standard in self.standards["binding"][sheetName]:
                self.add_binding_section(sheet, standard)
        fileName = self.saveDir + 'binding_template_' + receptor
        wb.save(fileName + '.xlsx')
        wb.close()
        self.add_macros(fileName, receptor)

    # Can produce function templates for 5HT or DA
    def make_function_template(self, receptor):
        match(receptor):
            case("DA"):
                wb = self.functionDA
                receptorList = self.receptorsDA
            case("5HT"):
                wb = self.function5HT
                receptorList = self.receptors5HT
        ws1 = wb.active
        ws1.title = receptorList[0] + ' agonist'
        wb.create_sheet(receptorList[0] + ' antagonist')
        for i in range(1, 4):
            wb.create_sheet(receptorList[i] + ' agonist')
            wb.create_sheet(receptorList[i] + ' antagonist')
        for sheet in receptorList:
            wsAgonist = wb[sheet + ' agonist']
            wsAgonist.append(self.sheetTemplate)
            wsAgonist.append(self.functionRow)
            wsAgonist.append([])
            self.format_cell(wsAgonist, 1, [1, 3, 5, 7], 'Title')
            self.format_cell(wsAgonist, 2, [1], 'Title')
            wsAntagonist = wb[sheet + ' antagonist']
            wsAntagonist.append(self.sheetTemplate)
            wsAntagonist.append(self.functionRow)
            wsAntagonist.append([])
            self.format_cell(wsAntagonist, 1, [1, 3, 5, 7], 'Title')
            self.format_cell(wsAntagonist, 2, [1], 'Title')
            startingRow = 4
            for i in self.drugNames:
                self.startRange = startingRow + 1
                self.endRange = startingRow + self.blankRows + 1
                wsAgonist.append(self.agonistTemplate)
                wsAgonist.cell(row=wsAgonist.max_row, column=1, value=i)
                self.format_cell(wsAgonist, startingRow, range(1, 16), 'Header')
                wsAntagonist.append(self.antagonistTemplate)
                wsAntagonist.cell(row=wsAntagonist.max_row, column=1, value=i)
                self.format_cell(wsAntagonist, startingRow, range(1, 16), 'Header')
                wsAgonist.append([])
                wsAgonist.cell(row=wsAgonist.max_row+1, column=8, 
                                value=self.add_formula('pctMax', column=None, row=wsAgonist.max_row+1))
                self.format_cell(wsAgonist, startingRow+1, [3, 6, 7], 'Table')
                wsAntagonist.append([])
                wsAntagonist.cell(row=wsAntagonist.max_row+1, column=8, 
                                value=self.add_formula('pctReversal', column=None, row=wsAntagonist.max_row+1))
                self.format_cell(wsAntagonist, startingRow+1, [3, 6, 7], 'Table')
                for i in range(0, self.blankRows-1):
                    wsAgonist.append([])
                    self.format_cell(wsAgonist, startingRow+2+i, [3, 6, 7], 'Table')
                    wsAntagonist.append([])
                    self.format_cell(wsAntagonist, startingRow+2+i, [3, 6, 7], 'Table')
                wsAgonist.append([])
                wsAntagonist.append([])
                self.populate_value_headers(wsAgonist, startingRow+1, ['C', 'H'])
                self.populate_value_headers(wsAntagonist, startingRow+1, ['C', 'H'])
                startingRow = startingRow + self.blankRows + 2
        fileName = self.saveDir + 'function_template_' + receptor
        wb.save(fileName + '.xlsx')
        wb.close()
        self.add_macros(fileName, receptor)

    # dictionary to store the formulae that populate the template worksheets in various combinations
    def add_formula(self, calculation, column=None, row=None):
        calculationDict = {
            'Mean': f'=AVERAGEIF({column}{self.startRange}:{column}{self.endRange}, \"<>0\")', 
            'SEM': f'=STDEV({column}{self.startRange}:{column}{self.endRange})/SQRT(COUNT({column}{self.startRange}:{column}{self.endRange}))', 
            'Ki': f'={column}{row}/(1+{self.ligandConcCol}{row}/{self.bindingKdCell})', 
            'pctMax': f'=(100-{self.maxEffectCol}{row})/(100-{self.standardCol}{row})*100',
            'pctReversal': f'=({self.maxEffectCol}{row}-{self.standardCol}{row})*100/(100-{self.standardCol}{row})',
            'Log': f'=LOG({column}{row}*0.000000001)'}
        formula = calculationDict[calculation]
        return formula

    # Places the appropriate mean and sem columns for a given worksheet
    def populate_value_headers(self, worksheet, row, columns):
        startingCell = ord('A')
        for i in columns:
            referenceCell = ord(i)
            meanCell = referenceCell - startingCell + 2
            semCell = referenceCell - startingCell + 3
            worksheet.cell(row=row, column=meanCell, 
                    value=self.add_formula('Mean', i))
            worksheet.cell(row=row, column = semCell, 
                    value=self.add_formula('SEM', i))
            
    def format_cell(self, sheet, row, columns, spec):

        def apply_formatting(cell, dict):
            for key, element in dict.items():
                exec(f"cell.{key} = element")
                
        defaultFont = "Arial"
        tableBorder = Side(border_style="thin", color="000000")
        headerBorder = Side(border_style="thick", color="000000")
        highlightBorder = Side(border_style="double", color="000000")
        formatDict = {
            'Title':{'font':Font(name=defaultFont, size=12, bold=True, color="000000")},
            'Header':{
                'font':Font(name=defaultFont, size=11, bold=True, color="000000"),
                'border':Border(left=tableBorder, right=tableBorder, top=headerBorder, bottom=headerBorder)
                },
            'Table':{
                'font':Font(name=defaultFont, size=11, bold=False, color="000000"),
                'border':Border(left=tableBorder, right=tableBorder)
                },
            'Highlight':{
                'font':Font(name="Arial", size=11, bold=False, color="FF0000"),
                'border':Border(left=highlightBorder, right=highlightBorder, top=highlightBorder, bottom=highlightBorder)
                }
        }
        currentFormat = formatDict[spec]
        for col in columns:
            currentCell = sheet.cell(row=row, column=col)
            apply_formatting(currentCell, currentFormat)

    def add_macros(self, fileName, receptor):
        match receptor:
            case("DA"):
                sheetNames = self.receptorsDA
            case("5HT"):
                sheetNames = self.receptors5HT
        vbaCode = """
            Private Sub Worksheet_Change(ByVal Target As Range)
                Dim ws As Worksheet
                Dim insertRow As Long
                Dim wasEmpty As Boolean
                Dim cell As Range

                Set ws = Me

                ' Avoid re-triggering the event
                Application.EnableEvents = False

                ' Only proceed if one cell was changed
                If Target.Cells.Count = 1 Then
                    insertRow = Target.Row

                    ' Check if the rest of the row was empty (excluding the changed cell)
                    wasEmpty = True
                    For Each cell In ws.Rows(insertRow).Cells
                        If cell.Address <> Target.Address And Not IsEmpty(cell.Value) Then
                            wasEmpty = False
                            Exit For
                        End If
                    Next cell

                    ' If the row was otherwise empty before this change, trigger the insert
                    If wasEmpty Then
                        ws.Rows(insertRow + 1).Insert Shift:=xlDown, CopyOrigin:=xlFormatFromLeftOrAbove
                    End If
                End If

                Application.EnableEvents = True
            End Sub
        """
        try:
            app = xw.App(visible=False)
            book = app.books.open(fileName + '.xlsx')
            book.save(fileName + '.xlsm')
            book.close()
            os.remove(fileName + '.xlsx')
            book = app.books.open(fileName + '.xlsm')

            for sheet in book.sheets:
                #sheetCodeName = book.api.Sheets(1).CodeName
                vbComponent = book.api.VBProject.VBComponents(sheet.api.CodeName)
                vbComponent.CodeModule.InsertLines(1, vbaCode)
            book.save()
            book.close()
        except Exception as e:
            print(f"Error: {e}")
            traceback.print_exc()
        finally:
            app.quit()

    def close_excel_processes(self):
        pass
