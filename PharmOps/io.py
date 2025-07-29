import pickle
import h5py
import pandas as pd
import numpy as np
from PharmOps import *
from importlib.metadata import version
from platformdirs import user_data_dir
import os
import re
from datetime import datetime
import openpyxl as pxl

# Parses a text file containing raw well-plate data and makes BindingPlate objects
def read_raw_well_txt(filepath: str, outputClass: WellData):
    rawData = pd.read_csv(filepath)
    pattern = r"Plate \d+.*?(?=Plate \d+|Total count rate:|$)"
    datePattern = r"\b(\d{1,2}-[A-Za-z]{3}-\d{4})\b"    # for regex to find the date
    for i in range(min(len(rawData), 10)):              # limit to the first 10 rows
        for col in rawData.columns:  
            cell = str(rawData.iloc[i][col])
            match = re.search(datePattern, cell)
            if match:
                assayDate = match.group(0)              # return the first matched date
                break
    # parse text file with regex to separate plates
    matches = re.findall(pattern, rawData.to_string(), re.DOTALL)
    plateData = {}
    rowLabels = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for match in matches:
        lines = re.split('\n', match)
        processedLines = []
        # clean escaped tabs
        for line in lines:
            line = line.replace('\\t', '\t')
            contents = line.split('\t')
            processedLines.append(contents)
        plateNumberMatch = re.search(r"plate (\d+)", match)
        if plateNumberMatch:
            plateNo = plateNumberMatch.group(1)
            plateData[f"Plate_{plateNo}"] = processedLines
    # make dataframe from cleaned text segments
    cleanedData = {}
    for plate, rows in plateData.items():
        numericRows = []
        for row in rows:
            rowLabel = row[0].strip()[-1]
            if rowLabel not in rowLabels:
                continue
            dataColumns = row[1:13]
            try:
                dataColumnsInt = [int(value) for value in dataColumns]
                numericRows.append(dataColumnsInt)
            except ValueError:
                continue
        df = pd.DataFrame(numericRows, columns=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])
        df.index = [row[0].strip()[-1] for row in rows if row[0].strip()[-1] in rowLabels]
        cleanedData[plate] = df
    # make BindingPlate array
    wellDataObjects = []
    for plate, df in cleanedData.items():
        wellDataObjects.append(outputClass(df, plate))
        wellDataObjects[-1].set_date(assayDate)
    return wellDataObjects

# Deprecated method for making BindingPlate from a csv
def read_raw_well_csv(filepath: str):
    rawData = pd.read_excel(filepath)
    rawWellValues = rawData.loc[rawData.index[0:8], range(1, 12)]    
    return rawWellValues

# uses os default data dir from installation to save h5 file
def get_default_h5_path(appName="PharmOps", fileName="data_store.h5"):
    # Determine the platform-specific data directory
    data_dir = user_data_dir(appName, appauthor=False)
    os.makedirs(data_dir, exist_ok=True)  # Ensure the directory exists
    return os.path.join(data_dir, fileName)

# turns date string into an appropriate format for h5 file
def convert_date_string(dateStr):
    parsedDate = datetime.strptime(dateStr, "%d-%b-%Y")
    parsedDate = parsedDate.strftime("%Y%m%d")
    return parsedDate

# sanitize input for pickling
def prepare_binary(obj):
    serializedObj = pickle.dumps(obj)
    serializedArray = np.frombuffer(serializedObj, dtype='uint8')
    return serializedArray

# save BindingPlate obj to h5
def save_new_BindingPlate(wellData: BindingPlate, filepath: str):
    serializedArray = prepare_binary(wellData)
    parsedDate = convert_date_string(wellData.metadata.date)
    plateNo = wellData.metadata.plateNo
    with h5py.File(filepath, "a") as h5file:
        group = h5file.require_group("assays/" + parsedDate)
        group.create_dataset(plateNo, data=serializedArray)
        group.attrs["version"] = version("PharmOps")

# save DrugReports obj to h5
def save_new_DrugReport(drugRep: DrugReports, filepath: str):
    serializedArray = prepare_binary(drugRep)
    drugName = drugRep.drug
    receptorName = drugRep.metadata.receptor
    parsedDate = convert_date_string(drugRep.metadata.date)
    with h5py.File(filepath, "a") as h5file:
        group = h5file.require_group("reports/" + drugName + "/" + receptorName)
        grp = group.create_dataset(parsedDate, data=serializedArray)
        group.attrs["version"] = version("PharmOps")

# save new screening for triplicates to h5
# this saves the actual plate stored by date hash
def save_new_screening_plate(plate: ScreeningPlate,
                             filepath: str):
    serializedArray = prepare_binary(plate)
    parsedDate = convert_date_string(plate.metadata.date)
    plateID = '+'.join([parsedDate, plate.metadata.plateNo])
    with h5py.File(filepath, "a") as h5file:
        plateGroup = h5file.require_group("screenings/" + "plates/" + plateID)
        tripletGroup = h5file.require_group("screenings/" + "triplicates/")
        for idx, metadata in plate.screeningDict.items():
            if metadata is None:
                continue
            drugName = metadata.drug
            receptor = metadata.receptor
            concentration = str(metadata.concentration)
            weighing = metadata.weighing
            plateHash = '+'.join([parsedDate, plate.metadata.plateNo, weighing])
            tripData = plate.get_triplicate_data(idx)
            tripIdx = str(idx)
            if type(metadata.concentration) is not int:
                tripContainer = tripletGroup.require_group(f"None/{receptor}/{metadata.concentration}/{plateHash}")
                tripContainer.create_dataset(tripIdx, data=tripData)
            else:
                tripContainer = tripletGroup.require_group(f"{drugName}/{receptor}/{concentration}/{plateHash}")
                tripContainer.create_dataset(tripIdx, data=tripData)
        plateGroup.create_dataset(plateID, data=serializedArray)

# this saves the drug receptor combo's reference to plate hash
def save_new_triplicate_assay(plate: ScreeningPlate, 
                              triplet: tuple, 
                              filepath: str):
    serializedArray = prepare_binary(triplet)

# returnes expDict for making a TriplicateScreen object   
def load_h5_triplicates(drugName, receptorName, concentration, h5file):
    triplicates = h5file['screenings']['triplicates']
    if drugName not in triplicates:
        raise KeyError(f"No drug '{drugName}' found in h5 file")
    if receptorName not in triplicates[drugName]:
        raise KeyError(f"No receptor '{receptorName}' data found for '{drugName}'")
    if concentration not in triplicates[drugName][receptorName]:
        raise KeyError(f"No assay at '{concentration}' concentration for drug '{drugName}' and receptor '{receptorName}'")
    allMatches = triplicates[drugName][receptorName][concentration]
    allKeys = list(allMatches.keys())
    experiments = {}
    for key in allKeys:
        if key not in experiments:
            experiments[key] = {'vals': [], 'nsb': [], 'totals': []}
        allTrips = allMatches[key].keys()
        for trip in list(allTrips):
            experiments[key]['vals'].append(allMatches[key][trip][:])
        nsb = triplicates["None"][receptorName]["Non Specific"]
        nsbTrips = nsb[key]
        for trip in list(nsbTrips):
            experiments[key]['nsb'].append(nsbTrips[trip][:])
        totals = triplicates["None"][receptorName]["Totals"]
        totalsTrips = totals[key]
        for trip in list(totalsTrips):
            experiments[key]['totals'].append(totalsTrips[trip][:])
    return experiments

# load existing DrugReports from h5 via unserializing 
def load_h5_DrugReports(drugName, receptorName, dateStr, filepath):
    with h5py.File(filepath, "r") as h5file:
        if drugName not in h5file["reports"]:
            raise KeyError(f"No drug '{drugName}' found in h5 file")
        if receptorName not in h5file["reports"][drugName]:
            raise KeyError(f"No receptor '{receptorName}' data found for '{drugName}'")
        if dateStr not in h5file["reports"][drugName][receptorName]:
            raise KeyError(f"No assay on '{dateStr}' for drug '{drugName}' and receptor '{receptorName}'")
        serializedArray = h5file["reports"][drugName][receptorName][dateStr][:]
    loadedRawData = pickle.loads(serializedArray.tobytes())
    return loadedRawData

# all of these methods below are deprecated and exist in bindMods.SummaryTable now
def find_excel_header(row, header):
    indices = [i for i, item in enumerate(row) if isinstance(item, str) and re.search(header, item, re.IGNORECASE)]
    return indices

# finds columns in excel file
def extract_mean_and_sem(row, identifier):
    averages = find_excel_header(row, "ave")
    sem = find_excel_header(row, "sem")
    idCol = find_excel_header(row, identifier)

def load_binding_summary_excel(filepath):
    allowedNames = ("5HT1A", "5HT2A", "5HT2B", "5HT2C", "D1", "D2", "D3", "D4.4")
    workbook = pxl.load_workbook(filepath, data_only=True)
    summaryDict = {}
    for name in workbook.sheetnames:
        matchingReceptor = [isinstance(item, str) and re.search(item, name, re.IGNORECASE) for item in allowedNames]
        print(name)
        if not any(matchingReceptor):
            print(name)
            continue
        summaryDict = parse_binding_summary_sheet(workbook, name, summaryDict)
    return summaryDict

# extracts data for summary table. currently a method for SummaryTable in bindMods
def parse_binding_summary_sheet(workbook, receptor, summary):
    sheet = workbook[receptor]
    keyHeaders = ['ic50', 'ki', 'hill slope']
    valHeaders = ['ave', 'sem']
    matchStr = r"\b(" + "|".join(map(re.escape, valHeaders)) + r")\b"
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        ic50 = find_excel_header(row, "ic50")
        ki = find_excel_header(row, "ki")
        hillSlope = find_excel_header(row, "hill slope")
        averages = find_excel_header(row, "ave")      # collected in case absolute refs are necessary
        sem = find_excel_header(row, "sem")           # collected in case absolute refs are necessary
        headerRow = find_excel_header(row, matchStr)
        if ic50 and ki and hillSlope:
            drugID = row[0]
            if not drugID:
                temp = idx
            while not drugID:
                temp = temp-1
                drugID = sheet.cell(row=temp+1, column=1).value
            # relative average and sem indexing to the cells with the datatype's header
            ic50Average = sheet.cell(row=idx+2, column=ic50[0]+2).value
            ic50SEM = sheet.cell(row=idx+2, column=ic50[0]+3).value
            kiAverage = sheet.cell(row=idx+2, column=ki[0]+2).value
            kiSEM = sheet.cell(row=idx+2, column=ki[0]+3).value
            hillAverage = sheet.cell(row=idx+2, column=hillSlope[0]+2).value
            hillSEM = sheet.cell(row=idx+2, column=hillSlope[0]+3).value
            summary[drugID][receptor]["mean"]["ic50"] = ic50Average
            summary[drugID][receptor]["sem"]["ic50"] = ic50SEM
            summary[drugID][receptor]["mean"]["ki"] = kiAverage
            summary[drugID][receptor]["sem"]["ki"] = kiSEM
            summary[drugID][receptor]["mean"]["hillSlope"] = hillAverage
            summary[drugID][receptor]["sem"]["hillSlope"] = hillSEM
    return summary

# currently a SummaryTable method
def make_function_table(srcDir):
    receptorNames = ("5HT1A", "5HT2A", "5HT2B", "5HTR", "5HT2C", "D1", "D2", "D3", "D4")
    dataIdentifiers =  ["FXN", "EC50"]
    localFiles = os.listdir(srcDir)
    fileId = {word: [file for file in localFiles if re.search(rf"\b{re.escape(word)}\b", file)] for word in receptorNames}
    print(fileId)
    for word in fileId:
        if not fileId[word]:
            continue
        for file in fileId[word]:
            fName, fExt = os.path.splitext(file)
            if fExt != ".xlsx":
                continue
            workbook = pxl.load_workbook(srcDir + "\\" + file, data_only=True)
            parse_wb_sheets(workbook)

# returns nothing?
def parse_wb_sheets(workbook, summary):
    summaryDict = {}
    for name in workbook.sheetnames:
        sheet = workbook[name]
        printNext = False
        for row in sheet.iter_rows(values_only=True):
            ec50 = find_excel_header(row, "ec50")
            if printNext:
                printNext = False
            if ec50:
                printNext = True

# for standards
def add_agency(filepath, agency):
    with h5py.File(filepath, 'a') as h5File:
        group = h5File.require_group("standards/" + agency)

# for standards
def add_receptor(filepath, agency, receptor):
    with h5py.File(filepath, 'a') as h5File:
        group = h5File.require_group("standards/" + agency + receptor)

# add new standards to h5
def add_drug_standards(filepath, agency, receptor, drug):
    with h5py.File(filepath, 'a') as h5File:
        group = h5File.require_group("standards/" + agency)
        dataset = group.create_dataset(receptor, data=drug)

# load standards from h5
def load_drug_standards(filepath, agency, receptor):
    with h5py.File(filepath, 'r') as h5File:
        if agency not in h5File["standards"]:
            raise KeyError(f"No data saved for '{agency}'")
        if receptor not in h5File["standards"][agency]:
            raise KeyError(f"No standards for receptor '{receptor}' for {agency}")
    savedStandards = h5File["standards"][agency][receptor][:]
